import sys
import os
import openpyxl
import json

def read_xlsx_to_json(file_path, output_path):
    print(f"Reading file: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = {}
    for name in wb.sheetnames:
        sheet = wb[name]
        sheet_rows = []
        for row in sheet.iter_rows(values_only=True):
            sheet_rows.append(list(row))
        data[name] = sheet_rows
        
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Written data to {output_path}")

if __name__ == '__main__':
    read_xlsx_to_json("catalog_products_2026-07-06 10_58_57.xlsx", "excel_contents.json")
